# requires: python-binance, functools, numpy, openpyxl
# To retrieve data from Binance we need an API key (read only enough)
# public and secret key should be placed in API.data (first and second line, just the keys)
## Corrections-commentaries-new ideas : matiasgcl@protonmail.com
## Was this script useful?
## spread the love: Ethereum erc20: 0xDc3d1a7566a536CFbcaAeb0CD2a179d78062B4b6

from binance.client import Client
from functools import reduce
import numpy as np
import time
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# retrieve top500 coins by MCap (CoinGecko)
import requests
cg_top250 = "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=250&page=1&sparkline=false"
cg_top2502 = "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=250&page=2&sparkline=false"
response = requests.get(cg_top250)
response2 = requests.get(cg_top2502)
top250 = response.json()
top2502 = response2.json()
symbol250 = [(i['symbol'].upper()) for i in top250]
symbol2502 = [(i['symbol'].upper()) for i in top2502]
symbol250.extend(symbol2502)
symbol250 = [s + 'USDT' for s in symbol250]

client = Client()

isUp = client.get_system_status()
assert isUp['status'] == 0 # Lets check we can actually retrieve information
print('System status ok\n')

info = client.get_all_tickers()
subs = 'USDT'
res = [i['symbol'] for i in info if subs in i['symbol']]

# now lets compare CG list with Binance USDT list
result = [x for x in symbol250 if x in res]
print('Number of coins meeting the criteria of being in CoinGecko\'s top 500 by marketcap and having a USDT pair on Binance: '+str(len(result)))
print('How many coins you want to list? ')
lim = input()

del result[int(lim):]

print('\nEnter timeframe for the computations.\n 1m, 3m, 5m, 15m, 30m, 1h, 2h, 4h, 6h, 8h, 12h, 1d, 3d, 1w, 1M\nWarning: Case Sensitive')
granularity = input()
print('\nEnter initial date-time (UTC) for data, format accepted is  <number> <month> <year> hh:mm:ss. \nExample: 01 May 2022 14:15:00\nIf time demanded is the open time for the corresponding candle, it will start with the next one available')
init_date = input()
print('\nEnter end date-time (UTC) for data, format accepted is  <number> <month> <year> hh:mm:ss. \nExample: 01 May 2022 14:15:00\nIf time demanded is the open time for the corresponding candle, it will start with the next one available')
end_date = input()

# example structure of kline data :
#[
#  [
#    1499040000000,      // Open time
#    "0.01634790",       // Open
#    "0.80000000",       // High
#    "0.01575800",       // Low
#    "0.01577100",       // Close
#    "148976.11427815",  // Volume
#    1499644799999,      // Close time
#    "2434.19055334",    // Quote asset volume
#    308,                // Number of trades
#    "1756.87402397",    // Taker buy base asset volume
#    "28.46694368",      // Taker buy quote asset volume
#    "17928899.62484339" // Ignore.
#  ]
#]

# Create workbook
wb = Workbook()

# add_sheet is used to create sheet.
FORMAT_CURRENCY_USD_SIMPLE_EXT = '"$"#,##0.0000_-'
sheet1 = wb.active
sheet1.title = 'Sorted by MCap - Spot USDT'
sheet1.cell(row=1, column=1, value='TimeFrame: '+granularity+' - Initial Date: '+init_date+' (UTC) - End Date: '+end_date+' (UTC)')
sheet1.merge_cells('A1:G1')
sheet1['A1'].alignment = Alignment(horizontal="center", vertical="center",shrink_to_fit=True)
sheet1.cell(row=2, column=1, value='Coin').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=2, value='Low').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=3, value='Time Low (Open)').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=4, value='High').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=5, value='Time High (Close)').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=6, value='Price Variation').alignment = Alignment(horizontal="center")
sheet1.cell(row=2, column=7, value='T low - T high').alignment = Alignment(horizontal="center")
data = []
perc = []
for i in range(len(result)):
    print('Processing Coin #'+str(i+1)+': '+result[i])
    klines = client.get_historical_klines(result[i], granularity, init_date, end_date)
    if(len(klines)>0):
        opentimes = [klines[n][0] for n in range(len(klines))]
        highs = [float(klines[n][2]) for n in range(len(klines))]
        lows = [float(klines[n][3]) for n in range(len(klines))]
        closetimes = [klines[n][6] for n in range(len(klines))]

        low = (min(lows))
        indexlow = lows.index(low)
        timelow = opentimes[indexlow]
        timelow_datetimeFormat = timelow/86400000+25569

        highsaux = highs[indexlow:]
        closetimesaux = closetimes[indexlow:]
        high = (max(highsaux))
        indexhigh = highsaux.index(high)
        timehigh = closetimesaux[indexhigh]
        timehigh_datetimeFormat = timehigh/86400000+25569

        var_perc = (float(high)-float(low))/float(low)
        timevar = (timehigh-timelow)/(1000*86400)

        sheet1.cell(row=i+3, column=1, value=result[i])
        sheet1.cell(row=i+3, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet1.cell(row=i+3, column=2, value=float(low))
        sheet1.cell(row=i+3, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE_EXT
        sheet1.cell(row=i+3, column=3, value=timelow_datetimeFormat)
        sheet1.cell(row=i+3, column=3).number_format = numbers.FORMAT_DATE_DATETIME
        sheet1.cell(row=i+3, column=4, value=float(high))
        sheet1.cell(row=i+3, column=4).number_format = FORMAT_CURRENCY_USD_SIMPLE_EXT
        sheet1.cell(row=i+3, column=5, value=timehigh_datetimeFormat)
        sheet1.cell(row=i+3, column=5).number_format = numbers.FORMAT_DATE_DATETIME
        sheet1.cell(row=i+3, column=6, value=var_perc)
        sheet1.cell(row=i+3, column=6).number_format = numbers.FORMAT_PERCENTAGE_00
        sheet1.cell(row=i+3, column=7, value=timevar)
        sheet1.cell(row=i+3, column=7).number_format = numbers.FORMAT_DATE_TIMEDELTA

        data.append([result[i],float(low),timelow_datetimeFormat,float(high),timehigh_datetimeFormat,var_perc,timevar])
        perc.append(var_perc)

    else:
        sheet1.cell(row=i+3, column=1, value=result[i]+': No data').alignment = Alignment(horizontal="center")
        data.append([result[i]+': No data','','','','',0,''])
        perc.append(0)
    time.sleep(0.05)

sheet1.column_dimensions['A'].width = 30
sheet1.column_dimensions['B'].width = 22
sheet1.column_dimensions['C'].width = 22
sheet1.column_dimensions['D'].width = 22
sheet1.column_dimensions['E'].width = 22
sheet1.column_dimensions['F'].width = 22
sheet1.column_dimensions['G'].width = 22
perc = np.multiply(-1,perc) # to inverse-sort (high to low)
sorted = np.argsort(perc)

sheet2 = wb.create_sheet("Sorted by Bounce - USDT Pairs", 0) # insert at first position
sheet2.cell(row=1, column=1, value='TimeFrame: '+granularity+' - Initial Date: '+init_date+' (UTC) - End Date: '+end_date+' (UTC)')
sheet2.merge_cells('A1:H1')
sheet2['A1'].alignment = Alignment(horizontal="center", vertical="center",shrink_to_fit=True)
sheet2.cell(row=2, column=1, value='Coin').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=2, value='Low').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=3, value='Time Low (Open)').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=4, value='High').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=5, value='Time High (Close)').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=6, value='Price Variation').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=7, value='T low - T high').alignment = Alignment(horizontal="center")
sheet2.cell(row=2, column=8, value='MCap ranking').alignment = Alignment(horizontal="center")

for i in range(len(data)):
    # data[i] = [result[i],float(low),timelow_datetimeFormat,float(high),timehigh_datetimeFormat,var_perc,timevar]
    sheet2.cell(row=i+3, column=1, value=data[sorted[i]][0])
    sheet2.cell(row=i+3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet2.cell(row=i+3, column=2, value=data[sorted[i]][1])
    sheet2.cell(row=i+3, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE_EXT
    sheet2.cell(row=i+3, column=3, value=data[sorted[i]][2])
    sheet2.cell(row=i+3, column=3).number_format = numbers.FORMAT_DATE_DATETIME
    sheet2.cell(row=i+3, column=4, value=data[sorted[i]][3])
    sheet2.cell(row=i+3, column=4).number_format = FORMAT_CURRENCY_USD_SIMPLE_EXT
    sheet2.cell(row=i+3, column=5, value=data[sorted[i]][4])
    sheet2.cell(row=i+3, column=5).number_format = numbers.FORMAT_DATE_DATETIME
    sheet2.cell(row=i+3, column=6, value=data[sorted[i]][5])
    sheet2.cell(row=i+3, column=6).number_format = numbers.FORMAT_PERCENTAGE_00
    sheet2.cell(row=i+3, column=7, value=data[sorted[i]][6])
    sheet2.cell(row=i+3, column=7).number_format = numbers.FORMAT_DATE_TIMEDELTA
    sheet2.cell(row=i+3, column=8, value=(sorted[i]+1))

sheet2.column_dimensions['A'].width = 30
sheet2.column_dimensions['B'].width = 22
sheet2.column_dimensions['C'].width = 22
sheet2.column_dimensions['D'].width = 22
sheet2.column_dimensions['E'].width = 22
sheet2.column_dimensions['F'].width = 22
sheet2.column_dimensions['G'].width = 22
sheet2.column_dimensions['H'].width = 22

saveas = 'Bounces-Top'+str(len(result))+'Coins-'+granularity+'-'+init_date+'-to-'+end_date+'.xls'
wb.save(filename=saveas)
print('Success! \nData is saved in file: '+saveas)
print('\nCorrections-commentaries-new ideas : matiasgcl@protonmail.com')
print('\nWas this script useful? spread the love: Ethereum erc20: 0xDc3d1a7566a536CFbcaAeb0CD2a179d78062B4b6')
